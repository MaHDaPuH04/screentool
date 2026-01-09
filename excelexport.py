import os
from datetime import datetime
from openpyxl import Workbook
from openpyxl.drawing.image import Image as ExcelImage
from config import config
from logger import logger


class ExcelExporter:
    def export_screenshots_to_excel(self, screenshots_folder, excel_path=None):
        """
        Экспортирует скриншоты по подпапкам с новыми именами групп.
        Пустые папки (без скриншотов) пропускаются.
        """
        print("=== НАЧАЛО ЭКСПОРТА ===")

        try:
            # Получаем все подпапки в правильном порядке
            subfolders = self._get_ordered_folders(screenshots_folder)
            
            print(f"Найдено подпапок: {len(subfolders)}")
            print(f"Все подпапки: {subfolders}")
            
            if not subfolders:
                return None, "Нет папок для экспорта"

            # Создаём Excel БЕЗ дефолтного листа
            workbook = Workbook()
            # СРАЗУ УДАЛЯЕМ ДЕФОЛТНЫЙ ЛИСТ
            if len(workbook.sheetnames) > 0:
                default_sheet = workbook[workbook.sheetnames[0]]
                workbook.remove(default_sheet)
                print("🗑️ Удален дефолтный лист 'Sheet1'")
            
            total_screens = 0
            created_sheets = 0

            for folder_name in subfolders:
                folder_path = os.path.join(screenshots_folder, folder_name)
                print(f"\n🔍 Проверяем папку: {folder_name}")
                
                # Получаем скриншоты из папки
                screenshots = self._get_screenshots_from_folder(folder_path)
                
                if not screenshots:
                    print(f"  ⚠️ Папка пуста, пропускаем")
                    continue
                
                # Создаем лист только если есть скриншоты
                sheet_name = folder_name[:31]  # Ограничение Excel на длину имени листа
                sheet = workbook.create_sheet(title=sheet_name)
                created_sheets += 1
                
                print(f"  ✅ Создан лист: '{sheet_name}' ({len(screenshots)} скриншотов)")
                
                # Добавляем скриншоты на лист
                added_count = self._add_screenshots_to_sheet(sheet, screenshots, folder_name)
                total_screens += added_count
            
            # Если вообще не создали ни одного листа
            if created_sheets == 0:
                print("\n❌ Нет скриншотов для экспорта (все папки пустые)")
                return None, "Нет скриншотов для экспорта"
            
            # Сохраняем файл
            if not excel_path:
                timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
                excel_path = os.path.join(screenshots_folder, f"Export_{timestamp}.xlsx")
            
            workbook.save(excel_path)
            
            print(f"\n✅ Успешно: {total_screens} скриншотов в {created_sheets} листах")
            print(f"✅ Файл: {os.path.basename(excel_path)}")
            
            logger.excel_export(excel_path, total_screens)
            
            return excel_path, f"Экспортировано {total_screens} скриншотов в {created_sheets} лист(ов)"
            
        except Exception as e:
            print(f"\n❌ Ошибка экспорта: {e}")
            import traceback
            traceback.print_exc()
            return None, f"Ошибка: {str(e)}"
    
    def _get_ordered_folders(self, base_path):
        """Возвращает папки в правильном порядке"""
        if not os.path.exists(base_path):
            return []
        
        # Предопределенный порядок
        predefined_order = ["poll+calib", "TIP", "ver", "TM", "PDT"]
        
        # Все существующие папки
        all_items = os.listdir(base_path)
        existing_folders = []
        
        # Сначала проверяем предопределенные в правильном порядке
        for folder_name in predefined_order:
            folder_path = os.path.join(base_path, folder_name)
            if os.path.isdir(folder_path):
                existing_folders.append(folder_name)
        
        # Затем добавляем press_ папки в правильном порядке
        press_folders = []
        for item in all_items:
            if item.startswith('press_'):
                folder_path = os.path.join(base_path, item)
                if os.path.isdir(folder_path):
                    try:
                        # Извлекаем номер
                        press_num = int(item.split('_')[1])
                        press_folders.append((press_num, item))
                    except (IndexError, ValueError):
                        press_folders.append((999, item))  # В конец если не распарсился
        
        # Сортируем press_ папки по номеру
        press_folders.sort(key=lambda x: x[0])
        for _, folder_name in press_folders:
            existing_folders.append(folder_name)
        
        return existing_folders
    
    def _get_screenshots_from_folder(self, folder_path):
        """Возвращает список скриншотов из папки"""
        if not os.path.exists(folder_path):
            return []
        
        screenshots = []
        try:
            # Ищем файлы скриншотов
            for filename in os.listdir(folder_path):
                full_path = os.path.join(folder_path, filename)
                
                # Проверяем что это файл и соответствует шаблону
                if os.path.isfile(full_path):
                    file_lower = filename.lower()
                    if (file_lower.endswith('.png') or 
                        file_lower.endswith('.jpg') or 
                        file_lower.endswith('.jpeg')):
                        if filename.startswith('screenshot_'):
                            screenshots.append(full_path)
        
        except Exception as e:
            print(f"  ⚠️ Ошибка чтения папки: {e}")
        
        # Сортируем по времени создания
        screenshots.sort(key=lambda x: os.path.getmtime(x))
        return screenshots
    
    def _add_screenshots_to_sheet(self, sheet, screenshots, folder_name):
        """Добавляет скриншоты на лист Excel"""
        current_row = 2
        added_count = 0
        
        for i, screenshot_path in enumerate(screenshots, 1):
            try:
                # Проверяем существование файла
                if not os.path.exists(screenshot_path):
                    print(f"    ⚠️ Файл не найден, пропускаем")
                    continue
                
                # Вставляем изображение
                img = ExcelImage(screenshot_path)
                img.anchor = f'A{current_row}'
                sheet.add_image(img)
                
                # Добавляем информацию
                filename = os.path.basename(screenshot_path)
                created_time = datetime.fromtimestamp(os.path.getctime(screenshot_path))
                
                sheet.cell(row=current_row, column=2, value=filename)
                sheet.cell(row=current_row, column=3, 
                          value=created_time.strftime('%Y-%m-%d %H:%M:%S'))
                sheet.cell(row=current_row, column=4, value=folder_name)
                
                # Вычисляем количество строк для изображения
                rows_needed = max(5, (img.height // 20) + 2)
                current_row += rows_needed
                added_count += 1
                
                print(f"    ✅ [{i}] {filename}")
                
            except Exception as e:
                print(f"    ❌ Ошибка {os.path.basename(screenshot_path)}: {e}")
                current_row += 10  # Пропускаем место при ошибке
        
        # Настраиваем ширину колонок если добавили скриншоты
        if added_count > 0:
            sheet.column_dimensions['A'].width = 30  # Изображения
            sheet.column_dimensions['B'].width = 25  # Имя файла
            sheet.column_dimensions['C'].width = 20  # Дата
            sheet.column_dimensions['D'].width = 15  # Группа
            
            # Добавляем заголовки
            sheet['B1'] = 'Имя файла'
            sheet['C1'] = 'Дата создания'
            sheet['D1'] = 'Группа'
            
            # Стили для заголовков (опционально)
            from openpyxl.styles import Font
            for cell in ['B1', 'C1', 'D1']:
                sheet[cell].font = Font(bold=True)
        
        return added_count